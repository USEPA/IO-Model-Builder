# This script converts the BEA 2002 make and use tables to Excel files that can
# be used in the data pipeline. The BEA 2002 use and make tables are provided in
#  a fixed size table format:

# use table:
# (0, 10),     # commodity IO code
# (10, 100),   # commodity description
# (100, 110),  # industry IO code
# (110, 200),  # industry description
# (200, 210)   # use table value

# make table:
# (0, 10),     # industry IO code
# (10, 100),   # industry description
# (100, 110),  # commodity IO code
# (110, 200),  # commodity description
# (200, 210)   # make table value

import openpyxl

# Column definitions of the BEA files
ROW_CODE = (0, 10)
ROW_NAME = (10, 100)
COL_CODE = (100, 110)
COL_NAME = (110, 200)
VALUE = (200, 210)

BEA_MAKE_TABLE = '../data/input/economic/bea2002/redef/IOMakeDetail.txt'
BEA_USE_TABLE = '../data/input/economic/bea2002/redef/IOUseDetail.txt'

OUT_INDUSTRIES = '../data/input/economic/industries.xlsx'
OUT_COMMODITIES = '../data/input/economic/commodities.xlsx'
OUT_MAKE_TABLE = '../data/input/economic/make_table.xlsx'
OUT_USE_TABLE = '../data/input/economic/use_table.xlsx'


def read_lines(file_path):
    with open(file_path, 'r', newline='\n') as f:
        i = -1
        for line in f:
            i += 1
            if i == 0:
                continue  # ignore header
            yield line


def get_field_value(field, line):
    f = line[field[0]:field[1]]
    return f.strip()


def each_entry(file_path):
    for line in read_lines(file_path):
        row_code = get_field_value(ROW_CODE, line)
        row_name = get_field_value(ROW_NAME, line)
        row_id = get_identifier(row_code, row_name)
        col_code = get_field_value(COL_CODE, line)
        col_name = get_field_value(COL_NAME, line)
        col_id = get_identifier(col_code, col_name)
        value = float(get_field_value(VALUE, line))
        yield (row_id, col_id, value)


def get_identifier(code, description):
    name = description
    if name.endswith('/1/'):
        name = name.replace('/1/', '').strip()
    return '%s - %s' % (code, name)


def make_industry_index():
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title='Industries', index=0)
    ws.cell(row=1, column=1, value='Industry identifier')
    industries = fetch_industries()
    for i in range(0, len(industries)):
        ws.cell(row=i+2, column=1, value=industries[i])
    wb.save(OUT_INDUSTRIES)


def fetch_industries():
    industries = []
    for entry in each_entry(BEA_MAKE_TABLE):
        industry = entry[0]
        if industry not in industries:
            industries.append(industry)
    for entry in each_entry(BEA_USE_TABLE):
        industry = entry[1]
        if industry not in industries:
            industries.append(industry)
    industries.sort()
    return industries


def make_commodity_index():
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title='Commodities', index=0)
    ws.cell(row=1, column=1, value='Commodity identifier')
    commodities = fetch_commodities()
    for i in range(0, len(commodities)):
        ws.cell(row=i+2, column=1, value=commodities[i])
    wb.save(OUT_COMMODITIES)


def fetch_commodities():
    commodities = []
    for entry in each_entry(BEA_MAKE_TABLE):
        commodity = entry[1]
        if commodity not in commodities:
            commodities.append(commodity)
    for entry in each_entry(BEA_USE_TABLE):
        commodity = entry[0]
        if commodity not in commodities:
            commodities.append(commodity)
    commodities.sort()
    return commodities


def make_make_table():
    industries = fetch_industries()
    ind_idx = {}
    commodities = fetch_commodities()
    com_idx = {}
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title='Make table', index=0)
    for i in range(0, len(industries)):
        ws.cell(row=i+2, column=1, value=industries[i])
        ind_idx[industries[i]] = i+2
    for i in range(0, len(commodities)):
        ws.cell(row=1, column=i+2, value=commodities[i])
        com_idx[commodities[i]] = i+2
    for entry in each_entry(BEA_MAKE_TABLE):
        row = ind_idx[entry[0]]
        col = com_idx[entry[1]]
        ws.cell(row=row, column=col, value=entry[2])
    wb.save(OUT_MAKE_TABLE)


def make_use_table():
    commodities = fetch_commodities()
    com_idx = {}
    industries = fetch_industries()
    ind_idx = {}
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title='Use table', index=0)
    for i in range(0, len(commodities)):
        ws.cell(row=i+2, column=1, value=commodities[i])
        com_idx[commodities[i]] = i+2
    for i in range(0, len(industries)):
        ws.cell(row=1, column=i+2, value=industries[i])
        ind_idx[industries[i]] = i+2
    for entry in each_entry(BEA_USE_TABLE):
        row = com_idx[entry[0]]
        col = ind_idx[entry[1]]
        ws.cell(row=row, column=col, value=entry[2])
    wb.save(OUT_USE_TABLE)


if __name__ == '__main__':
    make_industry_index()
    make_commodity_index()
    make_make_table()
    make_use_table()

